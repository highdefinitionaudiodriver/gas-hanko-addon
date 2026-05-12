# -*- coding: utf-8 -*-
"""
GAS Hanko Add-on 設計書生成スクリプト
ソースコードを解析し、design_document.xlsx を生成する
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================
#  共通スタイル定義
# ==========================================================
HEADER_FILL = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
HEADER_FONT = Font(name="Meiryo", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Meiryo", bold=True, size=14, color="B71C1C")
CELL_FONT = Font(name="Meiryo", size=10)
BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(ws, row, num_cols):
    """ヘッダー行のスタイルを設定"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_data_cells(ws, start_row, end_row, num_cols, alignment=TOP_LEFT):
    """データセルのスタイルを設定"""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = alignment
            cell.border = BORDER


def set_column_widths(ws, widths):
    """列幅を設定"""
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def add_title(ws, title, num_cols):
    """シートタイトルを設定"""
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.row_dimensions[1].height = 24


# ==========================================================
#  ワークブック作成
# ==========================================================
wb = Workbook()

# デフォルトシートを削除
wb.remove(wb.active)


# ----------------------------------------------------------
#  シート1: 機能一覧表（基本設計）
# ----------------------------------------------------------
ws1 = wb.create_sheet("1_機能一覧表")
add_title(ws1, "■ 機能一覧表（基本設計）— GAS Hanko Add-on v1.0.0", 6)

headers1 = ["機能ID", "機能名", "概要", "対象ユーザー", "対応アプリ", "実装ファイル"]
for i, h in enumerate(headers1, start=1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, len(headers1))

functions = [
    ["F-001", "メニュー追加", "ファイルを開いた際にメニューバーへ「電子印鑑」メニューを追加する。スプレッドシート/ドキュメント/スライドで共通動作。", "全ユーザー", "Sheets / Docs / Slides", "Code.gs (onOpen)"],
    ["F-002", "サイドバー表示", "メニューから呼び出され、印鑑生成用UIをサイドバーとして表示する。幅300px固定。", "全ユーザー", "Sheets / Docs / Slides", "Code.gs (showSidebar) / Sidebar.html"],
    ["F-003", "印鑑画像のリアルタイムプレビュー", "上段・中段・下段の入力に応じて、HTML5 Canvas上に日本式データ印を描画する。入力変更でリアルタイム更新。", "全ユーザー", "Sheets / Docs / Slides", "Sidebar.html (drawPreview)"],
    ["F-004", "今日の日付の自動入力", "サイドバー表示時に中段欄へ今日の日付（YYYY/MM/DD形式）を自動入力する。", "全ユーザー", "Sheets / Docs / Slides", "Sidebar.html (formatToday)"],
    ["F-005", "苗字の自動入力", "Googleアカウントのメールアドレスから苗字を取得し、上段欄へ自動入力する。マッピング未登録時は@より前の文字列を使用。", "全ユーザー", "Sheets / Docs / Slides", "Code.gs (getUserLastName)"],
    ["F-006", "トラッキングID生成", "押印ごとに8桁の16進ランダムIDを生成し、印影右下に極小グレー文字で埋め込む。簡易的な証跡管理。", "全ユーザー", "Sheets / Docs / Slides", "Sidebar.html (generateTrackingId)"],
    ["F-007", "スプレッドシート押印", "アクティブセルの位置に印鑑画像を60px正方形のOverGridImageとして挿入する。", "Sheetsユーザー", "Sheets", "Code.gs (insertIntoSpreadsheet_)"],
    ["F-008", "ドキュメント押印", "現在のカーソル位置にインライン画像（45pt正方形）として挿入する。カーソル取得失敗時は文書末尾に追加。", "Docsユーザー", "Docs", "Code.gs (insertIntoDocument_)"],
    ["F-009", "スライド押印", "現在選択中のスライドの中央に60pt正方形の画像として挿入する。スライド未選択時は1枚目にフォールバック。", "Slidesユーザー", "Slides", "Code.gs (insertIntoSlides_)"],
    ["F-010", "ホストアプリ自動判定", "実行コンテキストからホストアプリ（Sheets/Docs/Slides）を自動判定し、適切な処理に分岐する。", "システム内部", "Sheets / Docs / Slides", "Code.gs (getHostApp_)"],
    ["F-011", "ステータス表示", "押印成功時はトラッキングIDを含む成功メッセージ、失敗時はエラーメッセージをサイドバー下部に表示する。", "全ユーザー", "Sheets / Docs / Slides", "Sidebar.html (setStatus)"],
    ["F-012", "フォントサイズ自動調整", "印影の各段に入力された文字列が円内に収まるよう、フォントサイズを自動的に縮小する。最小8px。", "全ユーザー", "Sheets / Docs / Slides", "Sidebar.html (drawFittedText)"],
]

for r, row in enumerate(functions, start=4):
    for c, val in enumerate(row, start=1):
        ws1.cell(row=r, column=c, value=val)

style_data_cells(ws1, 4, 3 + len(functions), len(headers1))
set_column_widths(ws1, [10, 28, 60, 18, 25, 35])


# ----------------------------------------------------------
#  シート2: API仕様書（詳細設計）
# ----------------------------------------------------------
ws2 = wb.create_sheet("2_API仕様書")
add_title(ws2, "■ API仕様書（詳細設計）— GAS サーバー側関数仕様", 7)

ws2.cell(row=2, column=1, value="※ GASプロジェクトのため、本「API」は google.script.run 経由で呼ばれるサーバー側公開関数を指します。").font = Font(name="Meiryo", italic=True, size=9, color="666666")
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

headers2 = ["API ID", "関数名", "種別", "呼び出し方法", "引数", "戻り値", "処理概要"]
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, len(headers2))

apis = [
    ["API-001", "onOpen()", "トリガー", "シンプルトリガー（自動実行）", "なし", "void", "ファイル（Sheets/Docs/Slides）が開かれた際に自動実行され、メニューバーに「電子印鑑」メニューを追加する。getUi_() でホストアプリを判定。"],
    ["API-002", "showSidebar()", "メニュー", "メニュークリックにより呼び出し", "なし", "void", "Sidebar.html を読み込み、幅300pxのサイドバーとして表示する。タイトルは「電子印鑑（データ印）」。"],
    ["API-003", "getUserLastName()", "サーバー関数", "google.script.run.getUserLastName()", "なし", "String（苗字）", "Session.getActiveUser().getEmail() でメールアドレスを取得し、NAME_MAP に存在すれば対応する苗字、なければ @ より前の文字列を返す。メール取得失敗時は「山田」を返す。"],
    ["API-004", "insertStampImage(base64Data)", "サーバー関数", "google.script.run.insertStampImage(dataUrl)", "base64Data: String（data:image/png;base64,プレフィックス付き）", "String（成功メッセージ）", "Base64画像データをBlobに変換し、ホストアプリに応じた挿入関数（insertIntoSpreadsheet_/insertIntoDocument_/insertIntoSlides_）にディスパッチする。"],
    ["API-005", "getHostApp_()", "内部関数", "サーバー側内部呼び出し", "なし", "String（'spreadsheet'/'document'/'slides'/'unknown'）", "SpreadsheetApp/DocumentApp/SlidesApp の各 getActive*() を try-catch で順次試行し、成功したアプリ名を返す。プライベート関数（末尾アンダースコア）。"],
    ["API-006", "getUi_()", "内部関数", "サーバー側内部呼び出し", "なし", "Ui オブジェクト", "getHostApp_() の結果に基づき、対応するアプリの Ui オブジェクトを返す。未対応時は例外を投げる。"],
    ["API-007", "insertIntoSpreadsheet_(blob)", "内部関数", "サーバー側内部呼び出し", "blob: Blob（PNG画像）", "String（成功メッセージ）", "アクティブシートのアクティブセルの位置に sheet.insertImage() で画像を挿入し、60px正方形にリサイズする。"],
    ["API-008", "insertIntoDocument_(blob)", "内部関数", "サーバー側内部呼び出し", "blob: Blob（PNG画像）", "String（成功メッセージ）", "doc.getCursor() でカーソル位置を取得し、cursor.insertInlineImage() で挿入。45pt正方形にリサイズ。カーソル取得失敗時は body.appendParagraph() で末尾に追加。"],
    ["API-009", "insertIntoSlides_(blob)", "内部関数", "サーバー側内部呼び出し", "blob: Blob（PNG画像）", "String（成功メッセージ）", "presentation.getSelection().getCurrentPage() で選択中スライドを取得し、slide.insertImage() で挿入。60pt正方形にリサイズし、スライド中央に配置。"],
]

for r, row in enumerate(apis, start=5):
    for c, val in enumerate(row, start=1):
        ws2.cell(row=r, column=c, value=val)

style_data_cells(ws2, 5, 4 + len(apis), len(headers2))
set_column_widths(ws2, [10, 32, 12, 32, 35, 32, 60])

# クライアント側関数も追記
ws2.cell(row=5 + len(apis) + 1, column=1, value="■ クライアント側（Sidebar.html）主要関数").font = Font(name="Meiryo", bold=True, size=11, color="B71C1C")
client_header_row = 5 + len(apis) + 2
client_headers = ["関数ID", "関数名", "種別", "引数", "戻り値", "処理概要", "備考"]
for i, h in enumerate(client_headers, start=1):
    ws2.cell(row=client_header_row, column=i, value=h)
style_header_row(ws2, client_header_row, len(client_headers))

client_apis = [
    ["C-001", "drawPreview()", "プレビュー描画", "なし", "void", "現在の入力値で印影をプレビューCanvasに描画する。トラッキングIDなし。", "input イベントにバインド"],
    ["C-002", "drawStamp(c, size, top, mid, bottom, trackingId)", "印影描画", "Canvas Context, サイズ, 上段, 中段, 下段, トラッキングID", "void", "Canvas に円枠＋3分割線＋3段テキスト＋トラッキングIDを描画する。プレビュー/出力共通。", "コア描画関数"],
    ["C-003", "drawChordLine(c, cx, cy, r, y)", "弦描画", "Canvas Context, 中心X, 中心Y, 半径, 線のY座標", "void", "円の中で水平な弦（線分）を描画する。3段分割用。", "ピタゴラスの定理で弦長算出"],
    ["C-004", "drawFittedText(c, text, x, y, maxWidth, areaHeight, isDate)", "自動フィット文字描画", "Canvas Context, テキスト, X, Y, 最大幅, 領域高, 日付フラグ", "void", "テキストが領域に収まるまでフォントサイズを縮小しながら描画する。最小8px。", "日付は等幅フォント"],
    ["C-005", "generateTrackingId()", "ID生成", "なし", "String（8桁16進）", "crypto.getRandomValues() で4バイトの乱数を生成し、16進文字列として返す。", "Web Crypto API使用"],
    ["C-006", "formatToday()", "日付フォーマット", "なし", "String（YYYY/MM/DD）", "今日の日付をYYYY/MM/DD形式の文字列に整形して返す。", "ゼロパディング対応"],
    ["C-007", "onStamp()", "押印ハンドラ", "なし", "void", "オフスクリーンCanvasで印影を生成し、Base64 PNGに変換してサーバー側 insertStampImage() を呼び出す。", "押印ボタンクリック時"],
    ["C-008", "setStatus(msg, type)", "ステータス表示", "メッセージ, 種別（success/error/空）", "void", "サイドバー下部のステータス欄にメッセージを表示する。種別に応じてCSSクラスを切替。", ""],
]

for r, row in enumerate(client_apis, start=client_header_row + 1):
    for c, val in enumerate(row, start=1):
        ws2.cell(row=r, column=c, value=val)

style_data_cells(ws2, client_header_row + 1, client_header_row + len(client_apis), len(client_headers))


# ----------------------------------------------------------
#  シート3: テーブル定義書（詳細設計）
# ----------------------------------------------------------
ws3 = wb.create_sheet("3_テーブル定義書")
add_title(ws3, "■ テーブル定義書（詳細設計）— データ構造定義", 7)

ws3.cell(row=2, column=1, value="※ 本プロジェクトはGASベースであり、RDBは使用しません。代わりに、コード内で扱う主要なデータ構造（オブジェクト/定数）の定義を記載します。").font = Font(name="Meiryo", italic=True, size=9, color="666666")
ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

# データ構造1: NAME_MAP
ws3.cell(row=4, column=1, value="【テーブル1】 NAME_MAP（メールアドレス→苗字 マッピング）").font = Font(name="Meiryo", bold=True, size=11, color="B71C1C")
ws3.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)

ws3.cell(row=5, column=1, value="物理名: NAME_MAP").font = Font(name="Meiryo", size=10)
ws3.cell(row=5, column=3, value="型: Object（連想配列）").font = Font(name="Meiryo", size=10)
ws3.cell(row=5, column=5, value="格納場所: Code.gs (グローバル定数)").font = Font(name="Meiryo", size=10)
ws3.merge_cells(start_row=5, start_column=5, end_row=5, end_column=7)

headers3 = ["論理名", "物理名（キー）", "データ型", "主キー", "必須", "デフォルト", "説明"]
for i, h in enumerate(headers3, start=1):
    ws3.cell(row=6, column=i, value=h)
style_header_row(ws3, 6, len(headers3))

table1 = [
    ["メールアドレス", "(キー)", "String", "○", "○", "—", "Googleアカウントのメールアドレス（小文字推奨）。Session.getActiveUser().getEmail() の戻り値と完全一致で照合される。"],
    ["苗字", "(値)", "String", "—", "○", "—", "印影の上段に表示される日本語の苗字。最大6文字推奨。マッピングに登録されていれば自動入力される。"],
]

for r, row in enumerate(table1, start=7):
    for c, val in enumerate(row, start=1):
        ws3.cell(row=r, column=c, value=val)

style_data_cells(ws3, 7, 6 + len(table1), len(headers3))

# データ構造2: 印影描画パラメータ定数
ws3.cell(row=10, column=1, value="【テーブル2】 印影描画パラメータ（Sidebar.html JavaScript定数）").font = Font(name="Meiryo", bold=True, size=11, color="B71C1C")
ws3.merge_cells(start_row=10, start_column=1, end_row=10, end_column=7)

for i, h in enumerate(headers3, start=1):
    ws3.cell(row=11, column=i, value=h)
style_header_row(ws3, 11, len(headers3))

table2 = [
    ["印影サイズ", "STAMP_SIZE", "Number (px)", "—", "○", "200", "Canvas描画の解像度（幅・高さ共通、正方形）。プレビューと出力で共通使用。"],
    ["印鑑色", "STAMP_COLOR", "String (HEX)", "—", "○", "#c62828", "外枠・横線・テキスト全てに使用される赤色。"],
    ["トラッキングID色", "TRACKING_COLOR", "String (HEX)", "—", "○", "#aaaaaa", "印影右下に埋め込むトラッキングIDの文字色（グレー）。"],
    ["トラッキングID文字サイズ", "TRACKING_FONT_SIZE", "Number (px)", "—", "○", "8", "トラッキングIDの文字サイズ。極小表示。"],
]

for r, row in enumerate(table2, start=12):
    for c, val in enumerate(row, start=1):
        ws3.cell(row=r, column=c, value=val)

style_data_cells(ws3, 12, 11 + len(table2), len(headers3))

# データ構造3: 画像挿入サイズ定数
ws3.cell(row=17, column=1, value="【テーブル3】 画像挿入サイズ（Code.gs サーバー側変数）").font = Font(name="Meiryo", bold=True, size=11, color="B71C1C")
ws3.merge_cells(start_row=17, start_column=1, end_row=17, end_column=7)

for i, h in enumerate(headers3, start=1):
    ws3.cell(row=18, column=i, value=h)
style_header_row(ws3, 18, len(headers3))

table3 = [
    ["スプレッドシート挿入サイズ", "targetSize (Sheets)", "Number (px)", "—", "○", "60", "insertIntoSpreadsheet_() 内のローカル変数。OverGridImage の幅・高さ。"],
    ["ドキュメント挿入サイズ", "targetPt (Docs)", "Number (pt)", "—", "○", "45", "insertIntoDocument_() 内のローカル変数。InlineImage の幅・高さ。"],
    ["スライド挿入サイズ", "targetPt (Slides)", "Number (pt)", "—", "○", "60", "insertIntoSlides_() 内のローカル変数。スライド画像の幅・高さ。中央配置の計算にも使用。"],
]

for r, row in enumerate(table3, start=19):
    for c, val in enumerate(row, start=1):
        ws3.cell(row=r, column=c, value=val)

style_data_cells(ws3, 19, 18 + len(table3), len(headers3))

# データ構造4: フォーム入力項目
ws3.cell(row=23, column=1, value="【テーブル4】 サイドバー入力項目（Sidebar.html フォーム）").font = Font(name="Meiryo", bold=True, size=11, color="B71C1C")
ws3.merge_cells(start_row=23, start_column=1, end_row=23, end_column=7)

for i, h in enumerate(headers3, start=1):
    ws3.cell(row=24, column=i, value=h)
style_header_row(ws3, 24, len(headers3))

table4 = [
    ["上段テキスト", "top-text", "String", "—", "○", "（自動入力）", "印影の上段に表示。最大6文字。アカウント情報から自動入力される（API-003）。"],
    ["中段テキスト", "mid-text", "String", "—", "○", "今日の日付", "印影の中段に表示される日付。YYYY/MM/DD形式。最大10文字。等幅フォント。"],
    ["下段テキスト", "bottom-text", "String", "—", "—", "（空文字）", "印影の下段に表示。部署名等。最大6文字。空欄も可。"],
]

for r, row in enumerate(table4, start=25):
    for c, val in enumerate(row, start=1):
        ws3.cell(row=r, column=c, value=val)

style_data_cells(ws3, 25, 24 + len(table4), len(headers3))

set_column_widths(ws3, [22, 22, 14, 8, 8, 16, 50])


# ----------------------------------------------------------
#  シート4: エラー・ログ定義書（詳細設計）
# ----------------------------------------------------------
ws4 = wb.create_sheet("4_エラー・ログ定義書")
add_title(ws4, "■ エラー・ログ定義書（詳細設計）", 7)

headers4 = ["エラーID", "発生箇所", "エラー種別", "メッセージ", "発生条件", "ログレベル", "対処方法"]
for i, h in enumerate(headers4, start=1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, len(headers4))

errors = [
    ["E-001", "Code.gs / getUi_()", "ホスト判定エラー", "未対応のアプリケーションです", "Sheets/Docs/Slides 以外のコンテキストで onOpen() または showSidebar() が実行された場合", "ERROR", "対応アプリ（スプレッドシート、ドキュメント、スライド）から実行してください。"],
    ["E-002", "Code.gs / insertStampImage()", "ホスト判定エラー", "未対応のアプリケーションです: {host}", "getHostApp_() が 'unknown' を返した場合", "ERROR", "対応アプリから実行してください。スクリプトのコンテナ種別を確認してください。"],
    ["E-003", "Code.gs / insertIntoSlides_()", "スライド未存在エラー", "スライドが存在しません", "プレゼンテーション内に1枚もスライドがない場合（通常発生しない）", "ERROR", "プレゼンテーションに最低1枚のスライドを追加してください。"],
    ["E-004", "Code.gs / getUserLastName()", "メール取得失敗", "（例外なし、フォールバック動作）", "Session.getActiveUser().getEmail() が空文字を返す場合（共有アカウント等）", "WARN", "デフォルト値「山田」が返却されます。NAME_MAP の利用や手動入力で対応可能。"],
    ["E-005", "Code.gs / insertIntoDocument_()", "カーソル取得失敗", "（例外なし、フォールバック動作）", "ドキュメントで範囲選択中など、getCursor() が null を返す場合", "INFO", "ドキュメント末尾に画像が追加されます。「ドキュメント末尾に押印しました（カーソル位置を取得できなかったため）」のメッセージで通知。"],
    ["E-006", "Sidebar.html / onStamp() (failure)", "サーバー通信エラー", "エラー: {err.message}", "google.script.run の実行失敗時（権限不足、ネットワーク断、サーバー側例外など）", "ERROR", "ブラウザのネットワーク状態とGASプロジェクトの権限承認状態を確認してください。"],
    ["E-007", "ブラウザ / Web Crypto API", "Crypto API未サポート", "（例外発生）", "古いブラウザで crypto.getRandomValues() がサポートされていない場合", "ERROR", "最新版のChrome/Edge/Firefox/Safariを使用してください。"],
    ["E-008", "GAS実行環境", "OAuth権限未承認", "承認が必要です", "初回実行時、ユーザーがOAuth権限を承認していない場合", "WARN", "Apps Scriptエディタで初回実行を行い、表示される権限承認ダイアログで「許可」してください。"],
]

for r, row in enumerate(errors, start=4):
    for c, val in enumerate(row, start=1):
        ws4.cell(row=r, column=c, value=val)

style_data_cells(ws4, 4, 3 + len(errors), len(headers4))
set_column_widths(ws4, [10, 32, 18, 38, 45, 11, 50])

# ログ出力定義
log_start = 4 + len(errors) + 2
ws4.cell(row=log_start, column=1, value="■ ログ出力タイミング定義").font = Font(name="Meiryo", bold=True, size=11, color="B71C1C")
ws4.merge_cells(start_row=log_start, start_column=1, end_row=log_start, end_column=7)

log_headers = ["ログID", "出力箇所", "出力先", "出力タイミング", "出力内容", "ログレベル", "備考"]
for i, h in enumerate(log_headers, start=1):
    ws4.cell(row=log_start + 1, column=i, value=h)
style_header_row(ws4, log_start + 1, len(log_headers))

logs = [
    ["L-001", "サーバー側例外全般", "Stackdriver Logging（Google Cloud Logging）", "GAS実行中の未捕捉例外発生時", "スタックトレース、実行ユーザー、実行時刻", "ERROR", "appsscript.json の exceptionLogging: STACKDRIVER により自動記録"],
    ["L-002", "Sidebar.html / setStatus(success)", "サイドバー画面（DOM）", "押印成功時", "「{appName}に押印しました ({位置情報}) (ID: {trackingId})」", "INFO", "緑色テキストで表示"],
    ["L-003", "Sidebar.html / setStatus(error)", "サイドバー画面（DOM）", "押印失敗時", "「エラー: {err.message}」", "ERROR", "赤色テキストで表示"],
    ["L-004", "印影内", "印鑑画像内（右下）", "押印時", "8桁の16進トラッキングID", "INFO", "簡易証跡として印影に永続埋め込み"],
    ["L-005", "ブラウザ console.error", "ブラウザDevTools", "未捕捉JavaScriptエラー時", "エラーオブジェクト", "ERROR", "本実装では明示的な console.error 呼び出しなし"],
]

for r, row in enumerate(logs, start=log_start + 2):
    for c, val in enumerate(row, start=1):
        ws4.cell(row=r, column=c, value=val)

style_data_cells(ws4, log_start + 2, log_start + 1 + len(logs), len(log_headers))


# ----------------------------------------------------------
#  シート5: アーキテクチャ図解（Mermaidコード集）
# ----------------------------------------------------------
ws5 = wb.create_sheet("5_アーキテクチャ図解")
add_title(ws5, "■ アーキテクチャ図解（Mermaid記法）", 2)

ws5.cell(row=2, column=1, value="※ 各セルのMermaidコードを mermaid.live や VSCode の Mermaid プレビューで描画できます。").font = Font(name="Meiryo", italic=True, size=9, color="666666")
ws5.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

headers5 = ["図の名称", "Mermaidコード"]
for i, h in enumerate(headers5, start=1):
    ws5.cell(row=4, column=i, value=h)
style_header_row(ws5, 4, len(headers5))

# 図1: システム構成図
diagram1 = """graph TB
    subgraph "ユーザー環境"
        Browser["ブラウザ<br/>(Chrome/Edge/Safari/Firefox)"]
    end

    subgraph "Google Workspace"
        Sheets["Google スプレッドシート"]
        Docs["Google ドキュメント"]
        Slides["Google スライド"]
    end

    subgraph "Google Apps Script ランタイム"
        CodeGS["Code.gs<br/>(サーバー側ロジック)"]
        Sidebar["Sidebar.html<br/>(クライアント側UI)"]
        Manifest["appsscript.json<br/>(マニフェスト)"]
    end

    subgraph "Google API"
        SpreadsheetApp["SpreadsheetApp"]
        DocumentApp["DocumentApp"]
        SlidesApp["SlidesApp"]
        SessionAPI["Session API<br/>(getActiveUser)"]
    end

    Browser -->|操作| Sheets
    Browser -->|操作| Docs
    Browser -->|操作| Slides

    Sheets -->|onOpen| CodeGS
    Docs -->|onOpen| CodeGS
    Slides -->|onOpen| CodeGS

    CodeGS -->|HtmlService| Sidebar
    Sidebar -->|google.script.run| CodeGS

    CodeGS --> SpreadsheetApp
    CodeGS --> DocumentApp
    CodeGS --> SlidesApp
    CodeGS --> SessionAPI

    SpreadsheetApp --> Sheets
    DocumentApp --> Docs
    SlidesApp --> Slides"""

# 図2: 画面遷移図
diagram2 = """stateDiagram-v2
    [*] --> ファイルオープン
    ファイルオープン --> メニュー表示: onOpen()自動実行
    メニュー表示 --> サイドバー表示: 「電子印鑑」→「印鑑パネルを開く」クリック

    state サイドバー表示 {
        [*] --> 初期化
        初期化 --> 苗字取得中: getUserLastName呼出
        苗字取得中 --> 入力待機: 自動入力完了
        入力待機 --> 入力中: ユーザー入力
        入力中 --> 入力待機: プレビュー更新
        入力待機 --> 押印処理中: 押印ボタンクリック
        押印処理中 --> 成功表示: insertStampImage成功
        押印処理中 --> エラー表示: insertStampImage失敗
        成功表示 --> 入力待機
        エラー表示 --> 入力待機
    }

    サイドバー表示 --> ファイルオープン: サイドバーを閉じる"""

# 図3: ER図相当（データ構造図）
diagram3 = """erDiagram
    NAME_MAP {
        string email PK "メールアドレス（キー）"
        string lastName "苗字（値）"
    }

    SIDEBAR_INPUT {
        string topText "上段テキスト（最大6文字）"
        string midText "中段テキスト（YYYY/MM/DD）"
        string bottomText "下段テキスト（最大6文字）"
    }

    STAMP_IMAGE {
        string base64Data "data:image/png;base64,..."
        string trackingId "8桁16進ID"
        number size "200px"
        string color "#c62828"
    }

    INSERTED_IMAGE {
        string hostApp "spreadsheet/document/slides"
        number width "60px or 45pt or 60pt"
        number height "60px or 45pt or 60pt"
        string position "セル/カーソル/スライド中央"
    }

    NAME_MAP ||--o| SIDEBAR_INPUT : "topTextに自動入力"
    SIDEBAR_INPUT ||--|| STAMP_IMAGE : "Canvas描画"
    STAMP_IMAGE ||--|| INSERTED_IMAGE : "Base64→Blob変換後挿入" """

# 図4: シーケンス図（押印処理の全体フロー）
diagram4 = """sequenceDiagram
    actor User as ユーザー
    participant Browser as ブラウザ
    participant Sidebar as Sidebar.html
    participant Canvas as HTML5 Canvas
    participant CodeGS as Code.gs
    participant GAS as Google Apps Script
    participant App as Sheets/Docs/Slides

    User->>Browser: ファイルを開く
    Browser->>App: ファイル表示
    App->>CodeGS: onOpen()トリガー
    CodeGS->>CodeGS: getHostApp_()
    CodeGS->>App: メニュー追加

    User->>App: 「電子印鑑」→「印鑑パネルを開く」
    App->>CodeGS: showSidebar()
    CodeGS->>Sidebar: HtmlService.createHtmlOutputFromFile
    Sidebar->>Browser: サイドバー表示

    Sidebar->>CodeGS: getUserLastName() (google.script.run)
    CodeGS->>GAS: Session.getActiveUser().getEmail()
    GAS-->>CodeGS: email
    CodeGS->>CodeGS: NAME_MAP[email] 検索
    CodeGS-->>Sidebar: 苗字
    Sidebar->>Sidebar: top-text に自動入力

    User->>Sidebar: 各欄を入力
    Sidebar->>Canvas: drawPreview() リアルタイム描画

    User->>Sidebar: 押印ボタンクリック
    Sidebar->>Sidebar: generateTrackingId()
    Sidebar->>Canvas: drawStamp(offCanvas, ..., trackingId)
    Canvas-->>Sidebar: PNG dataURL

    Sidebar->>CodeGS: insertStampImage(dataUrl) (google.script.run)
    CodeGS->>CodeGS: Base64デコード→Blob化
    CodeGS->>CodeGS: getHostApp_()

    alt スプレッドシート
        CodeGS->>App: sheet.insertImage(blob, col, row)
    else ドキュメント
        CodeGS->>App: cursor.insertInlineImage(blob)
    else スライド
        CodeGS->>App: slide.insertImage(blob)
    end

    App-->>CodeGS: 挿入完了
    CodeGS-->>Sidebar: 成功メッセージ
    Sidebar->>Browser: ステータス表示（成功）"""

# 図5: クラス図（モジュール構成）
diagram5 = """classDiagram
    class CodeGS {
        +NAME_MAP: Object
        +onOpen() void
        +showSidebar() void
        +getUserLastName() String
        +insertStampImage(base64Data) String
        -getHostApp_() String
        -getUi_() Ui
        -insertIntoSpreadsheet_(blob) String
        -insertIntoDocument_(blob) String
        -insertIntoSlides_(blob) String
    }

    class SidebarHTML {
        +STAMP_SIZE: Number
        +STAMP_COLOR: String
        +TRACKING_COLOR: String
        +TRACKING_FONT_SIZE: Number
        +drawPreview() void
        +drawStamp(c, size, top, mid, bottom, trackingId) void
        +drawChordLine(c, cx, cy, r, y) void
        +drawFittedText(c, text, x, y, maxWidth, areaHeight, isDate) void
        +generateTrackingId() String
        +formatToday() String
        +onStamp() void
        +setStatus(msg, type) void
    }

    class AppsScriptJSON {
        +timeZone: String
        +runtimeVersion: String
        +oauthScopes: Array
    }

    class GoogleAPIs {
        <<external>>
        +SpreadsheetApp
        +DocumentApp
        +SlidesApp
        +Session
        +HtmlService
        +Utilities
    }

    CodeGS --> GoogleAPIs : 使用
    CodeGS --> SidebarHTML : HtmlService.createHtmlOutputFromFile
    SidebarHTML --> CodeGS : google.script.run
    AppsScriptJSON ..> CodeGS : 設定"""

# 図6: フローチャート（ホストアプリ判定ロジック）
diagram6 = """flowchart TD
    Start([insertStampImage呼出]) --> Decode[Base64デコード→Blob化]
    Decode --> GetHost[getHostApp_実行]

    GetHost --> TrySheets{SpreadsheetApp<br/>取得試行}
    TrySheets -->|成功| Sheets[insertIntoSpreadsheet_]
    TrySheets -->|失敗| TryDocs{DocumentApp<br/>取得試行}

    TryDocs -->|成功| Docs[insertIntoDocument_]
    TryDocs -->|失敗| TrySlides{SlidesApp<br/>取得試行}

    TrySlides -->|成功| Slides[insertIntoSlides_]
    TrySlides -->|失敗| Error[未対応エラー]

    Sheets --> InsCell[アクティブセル位置に<br/>60px画像挿入]
    Docs --> CheckCursor{カーソル取得}
    CheckCursor -->|成功| InsCursor[カーソル位置に<br/>45ptインライン画像挿入]
    CheckCursor -->|失敗| InsBody[文書末尾に追加]

    Slides --> GetSelection[選択中スライド取得]
    GetSelection --> InsSlide[スライド中央に<br/>60pt画像挿入]

    InsCell --> Success([成功メッセージ返却])
    InsCursor --> Success
    InsBody --> Success
    InsSlide --> Success
    Error --> Fail([例外スロー])"""

diagrams = [
    ["1. システム構成図", diagram1],
    ["2. 画面遷移図", diagram2],
    ["3. ER図（データ構造関連図）", diagram3],
    ["4. シーケンス図（押印処理フロー）", diagram4],
    ["5. クラス図（モジュール構成）", diagram5],
    ["6. フローチャート（ホストアプリ判定ロジック）", diagram6],
]

for r, (name, code) in enumerate(diagrams, start=5):
    ws5.cell(row=r, column=1, value=name)
    ws5.cell(row=r, column=2, value=code)

style_data_cells(ws5, 5, 4 + len(diagrams), 2)

# Mermaidコード列は等幅フォントにし、各行を高くする
mono_font = Font(name="Consolas", size=9)
for r in range(5, 5 + len(diagrams)):
    ws5.cell(row=r, column=1).font = Font(name="Meiryo", bold=True, size=10)
    ws5.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws5.cell(row=r, column=2).font = mono_font
    ws5.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    # 行高を内容に応じて調整
    line_count = diagrams[r - 5][1].count("\n") + 1
    ws5.row_dimensions[r].height = min(line_count * 14, 600)

set_column_widths(ws5, [35, 130])


# ----------------------------------------------------------
#  シート0: 表紙
# ----------------------------------------------------------
ws0 = wb.create_sheet("0_表紙", 0)
ws0.cell(row=2, column=2, value="GAS Hanko Add-on").font = Font(name="Meiryo", bold=True, size=24, color="B71C1C")
ws0.cell(row=3, column=2, value="設計書 (Design Document)").font = Font(name="Meiryo", bold=True, size=18, color="333333")
ws0.cell(row=5, column=2, value="バージョン: 1.0.0").font = Font(name="Meiryo", size=12)
ws0.cell(row=6, column=2, value="作成日: 2026-04-06").font = Font(name="Meiryo", size=12)
ws0.cell(row=7, column=2, value="対象アプリ: Google スプレッドシート / ドキュメント / スライド").font = Font(name="Meiryo", size=12)
ws0.cell(row=8, column=2, value="ランタイム: Google Apps Script V8").font = Font(name="Meiryo", size=12)

ws0.cell(row=11, column=2, value="■ シート構成").font = Font(name="Meiryo", bold=True, size=14, color="B71C1C")

toc = [
    ["シート名", "内容"],
    ["1_機能一覧表", "システムが提供する機能のID、機能名、概要、対象ユーザー、対応アプリ、実装ファイルの一覧（基本設計）"],
    ["2_API仕様書", "サーバー側公開関数および内部関数、クライアント側関数の仕様詳細（詳細設計）"],
    ["3_テーブル定義書", "コード内で扱うデータ構造（NAME_MAP、定数、フォーム入力項目）の定義（詳細設計）"],
    ["4_エラー・ログ定義書", "システム内で発生するエラーとログ出力の定義（詳細設計）"],
    ["5_アーキテクチャ図解", "システム構成図、画面遷移図、ER図、シーケンス図、クラス図、フローチャートをMermaid記法で記述"],
]

for r, row in enumerate(toc, start=12):
    for c, val in enumerate(row, start=2):
        cell = ws0.cell(row=r, column=c, value=val)
        if r == 12:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        else:
            cell.font = CELL_FONT
        cell.alignment = LEFT if c == 3 else CENTER
        cell.border = BORDER

ws0.cell(row=20, column=2, value="■ プロジェクト情報").font = Font(name="Meiryo", bold=True, size=14, color="B71C1C")
info = [
    ["プロジェクト名", "GAS Hanko Add-on"],
    ["概要", "日本のビジネスシーンで使われるデータ印（電子印鑑）をGoogle Workspaceの各アプリに挿入するアドオン"],
    ["ライセンス", "MIT License"],
    ["作者", "highdefinitionaudiodriver"],
    ["主要技術", "Google Apps Script V8 / HTML Service / HTML5 Canvas / Web Crypto API"],
    ["ファイル数", "3 (Code.gs / Sidebar.html / appsscript.json)"],
]

for r, row in enumerate(info, start=21):
    for c, val in enumerate(row, start=2):
        cell = ws0.cell(row=r, column=c, value=val)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER
    ws0.cell(row=r, column=2).font = Font(name="Meiryo", bold=True, size=10)

set_column_widths(ws0, [3, 22, 90])


# ==========================================================
#  保存
# ==========================================================
output_path = "G:/マイドライブ/claudecode/gas-hanko-addon/design_document.xlsx"
wb.save(output_path)
print(f"設計書を生成しました: {output_path}")
print(f"シート数: {len(wb.sheetnames)}")
print(f"シート一覧: {wb.sheetnames}")
